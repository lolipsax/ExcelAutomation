#!/usr/bin/env python3
"""Inspect the Excel workbooks in this repository.

Walks the repo for ``.xlsx`` files and prints a summary for each one
(sheet names, dimensions and header row). This doubles as a smoke test
that the development environment can actually read the project's data.

Usage:
    python scripts/inspect_workbooks.py [root_dir]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def iter_workbooks(root: Path):
    for path in sorted(root.rglob("*.xlsx")):
        # Skip Excel lock/temp files (e.g. "~$Book.xlsx").
        if path.name.startswith("~$"):
            continue
        yield path


def summarize(path: Path) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"\n{path}")
    for ws in wb.worksheets:
        header = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header = [str(c) for c in row if c is not None][:6]
            break
        print(
            f"  - sheet {ws.title!r}: {ws.max_row} rows x {ws.max_column} cols"
            + (f" | header: {header}" if header else "")
        )
    wb.close()


def main() -> int:
    root = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.cwd()
    workbooks = list(iter_workbooks(root))
    if not workbooks:
        print(f"No .xlsx files found under {root}")
        return 1

    print(f"Found {len(workbooks)} workbook(s) under {root}")
    for path in workbooks:
        try:
            summarize(path)
        except Exception as exc:  # noqa: BLE001 - report and continue
            print(f"\n{path}\n  ! failed to read: {exc}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
